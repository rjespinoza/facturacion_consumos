# import streamlit as st
# import polars as pl
# from openpyxl import Workbook
# import bcrypt
# import io
# import os
# import tempfile
# import re

# # Función para verificar usuario y contraseña
# def check_password():
#     """Función para autenticar al usuario."""
#     def password_entered():
#         if bcrypt.checkpw(st.session_state["password"].encode("utf-8"), PASSWORD_CORRECTO):
#             st.session_state["password_correct"] = True
#         else:
#             st.session_state["password_correct"] = False
#             st.error("Usuario/contraseña incorrectos")

#     if "password_correct" not in st.session_state or not st.session_state["password_correct"]:
#         st.text_input("Usuario", key="username")
#         st.text_input("Contraseña", type="password", key="password", on_change=password_entered)
#         st.stop()
    
#     return True

# # Usuario y contraseña predefinidos
# USUARIO_CORRECTO = "admin"
# PASSWORD_CORRECTO = bcrypt.hashpw("admin".encode("utf-8"), bcrypt.gensalt())

# # Interfaz Principal
# st.title("Extracción y filtrado de datos XLSB (Mes/Año)")

# if check_password():
#     uploaded_file = st.file_uploader("Cargar archivo XLSB", type="xlsb")

#     if uploaded_file is not None:
#         try:
#             # Configuración fija
#             HOJA_FIJA = 'Libro Fac.Emitidas MM 2025'
#             COLUMNAS_FIJAS = ['CUPS', 'Tipo factura', 'Fecha factura', 'Fecha inicial', 'Fecha final', 'Num. días', 'Tarifa', 'Total término variable (kWh)']

#             # Entrada de Año-Mes con validación
#             AÑO_MES_FILTRO = st.text_input("Año-Mes (YYYY-MM)", "2025-01")

#             if not re.match(r"^\d{4}-\d{2}$", AÑO_MES_FILTRO):
#                 st.error("Formato incorrecto. Use el formato YYYY-MM (ej. 2025-01)")
#                 st.stop()

#             # Guardar archivo temporalmente
#             with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
#                 tmp_file.write(uploaded_file.read())
#                 tmp_file_path = tmp_file.name

#             # Leer archivo XLSB con Polars
#             try:
#                 df = pl.read_excel(tmp_file_path, sheet_name=HOJA_FIJA).select(COLUMNAS_FIJAS)
#             except Exception as e:
#                 st.error(f"Error al leer la hoja '{HOJA_FIJA}'. Verifique el nombre de la hoja.")
#                 st.stop()

#             # Mostrar los primeros datos para depuración
#             st.write("Vista previa de los datos cargados:")
#             st.dataframe(df.head())

#             # Opción para eliminar datos nulos
#             if st.checkbox("Eliminar datos nulos"):
#                 df = df.drop_nulls()

#             # Validar existencia de la columna antes de convertir
#             if "Fecha inicial" not in df.columns:
#                 st.error("La columna 'Fecha inicial' no existe en los datos.")
#                 st.stop()

#             # Convertir 'Fecha inicial' correctamente a tipo fecha
#             df = df.with_columns(
#                 pl.col("Fecha inicial").cast(pl.Date)
#             )

#             # Filtrado por Año-Mes
#             df_filtered = df.filter(
#                 (pl.col("Tipo factura") == "Normal") &
#                 (pl.col("Fecha inicial").dt.strftime("%Y-%m") == AÑO_MES_FILTRO)
#             )

#             # Mostrar datos filtrados
#             if not df_filtered.is_empty():
#                 st.write("Datos filtrados:")
#                 st.dataframe(df_filtered)
#             else:
#                 st.warning("No se encontraron datos para el Año-Mes seleccionado.")

#             # Opción de descarga
#             output = io.BytesIO()
#             wb_out = Workbook()
#             ws = wb_out.active
#             ws.append(list(df_filtered.columns))
#             for row in df_filtered.to_numpy():
#                 ws.append(list(row))
#             wb_out.save(output)
#             output.seek(0)

#             st.download_button(
#                 label="Descargar datos filtrados (Excel)",
#                 data=output,
#                 file_name="datos_filtrados.xlsx",
#                 mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )

#         except Exception as e:
#             st.error(f"Error inesperado: {e}")

#         finally:
#             if 'tmp_file_path' in locals() and os.path.exists(tmp_file_path):
#                 os.remove(tmp_file_path)

# -------------------------------------------------
# import streamlit as st
# import polars as pl
# from openpyxl import Workbook
# import bcrypt
# import io
# import os
# import tempfile
# import re

# # Función para verificar el usuario y contraseña
# def check_password():
#     """Función para autenticar al usuario."""
#     def password_entered():
#         if bcrypt.checkpw(st.session_state["username"].encode("utf-8"), PASSWORD_CORRECTO):
#             st.session_state["password_correct"] = True
#             del st.session_state["password"]  # No almacenar la contraseña en la sesión
#         else:
#             st.session_state["password_correct"] = False

#     if "password_correct" not in st.session_state:
#         st.text_input("Usuario", key="username")
#         st.text_input("Contraseña", type="password", key="password", on_change=password_entered)
#         return False
#     elif not st.session_state["password_correct"]:
#         st.text_input("Usuario", key="username")
#         st.text_input("Contraseña", type="password", key="password", on_change=password_entered)
#         st.error("Usuario/contraseña incorrectos")
#         return False
#     else:
#         return True

# # Datos de usuario (¡Importante! Cambia esto por un sistema de almacenamiento seguro)
# USUARIO_CORRECTO = "admin"
# PASSWORD_CORRECTO = bcrypt.hashpw("admin".encode("utf-8"), bcrypt.gensalt())

# # Interfaz Principal
# st.title("Extracción y filtrado de datos XLSB (Mes/Año)")

# if check_password():
#     uploaded_file = st.file_uploader("Cargar archivo XLSB", type="xlsb")

#     if uploaded_file is not None:
#         try:
#             # Configuración fija
#             HOJA_FIJA = 'Libro Fac.Emitidas MM 2025'
#             COLUMNAS_FIJAS = ['CUPS', 'Tipo factura', 'Fecha factura', 'Fecha inicial', 'Fecha final', 'Num. días', 'Tarifa', 'Total término variable (kWh)']
#             AÑO_MES_FILTRO = st.text_input("Año-Mes (YYYY-MM)", "2025-01")

#             # Validar el formato de la entrada del usuario
#             if not re.match(r"^\d{4}-\d{2}$", AÑO_MES_FILTRO):
#                 st.error("Formato de Año-Mes incorrecto (YYYY-MM)")
#                 st.stop()

#             with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
#                 tmp_file.write(uploaded_file.read())
#                 tmp_file_path = tmp_file.name

#             df = pl.read_excel(tmp_file_path, sheet_name=HOJA_FIJA).select(COLUMNAS_FIJAS)

#             if st.checkbox("Eliminar datos nulos"):
#                 df = df.drop_nulls()

#             df_filtered = df.filter(
#                 (pl.col("Tipo factura") == "Normal") &
#                 (pl.col("Fecha inicial").cast(pl.Utf8).str.slice(0, 7) == AÑO_MES_FILTRO)
#             )

#             # Copia antes de eliminar la columna 'Tipo factura'
#             df_consumos = df_filtered.clone()

#             # Mostrar datos en la app
#             st.dataframe(df_filtered)

#             # Generar tablas pivoteadas acumulando por mes
#             df_grouped = df_filtered.with_columns(
#                 pl.col('Fecha inicial').dt.strftime('%Y-%m').alias('Mes')
#             ).group_by(['Tarifa', 'Mes']).agg([
#                 pl.col('Num. días').sum().alias('Suma de num. dias'),
#                 pl.col('CUPS').count().alias('Cuenta de CUPS'),
#                 pl.col('Total término variable (kWh)').sum().alias('Suma de kWh')
#             ])

#             df_pivot_num_dias = df_grouped.pivot(values='Suma de num. dias', index='Tarifa', on='Mes', sort_columns=True)
#             df_pivot_cuenta_cups = df_grouped.pivot(values='Cuenta de CUPS', index='Tarifa', on='Mes', sort_columns=True)
#             df_pivot_suma_kwh = df_grouped.pivot(values='Suma de kWh', index='Tarifa', on='Mes', sort_columns=True)

#             # Guardar los datos en Excel
#             output = io.BytesIO()
#             wb_out = Workbook()

#             # Hoja 'raw_data' con los datos filtrados
#             ws_raw_data = wb_out.active
#             ws_raw_data.title = "raw_data"
#             ws_raw_data.append(list(df_filtered.columns))
#             for row in df_filtered.to_numpy():
#                 ws_raw_data.append(list(row))

#             # Hoja 'consumos' con las tablas pivot
#             ws_consumos = wb_out.create_sheet(title="consumos")

#             def write_dataframe(ws, df, start_row):
#                 """Función auxiliar para escribir un DataFrame en la hoja de Excel."""
#                 ws.append([])
#                 for row in df.to_pandas().itertuples(index=False):
#                     ws.append(row)
#                 return start_row + len(df) + 2  # Deja una fila vacía

#             start_row = 1
#             ws_consumos.append(["Tabla: Suma de num. dias"])
#             start_row = write_dataframe(ws_consumos, df_pivot_num_dias, start_row)

#             ws_consumos.append(["Tabla: Cuenta de CUPS"])
#             start_row = write_dataframe(ws_consumos, df_pivot_cuenta_cups, start_row)

#             ws_consumos.append(["Tabla: Suma de kWh"])
#             start_row = write_dataframe(ws_consumos, df_pivot_suma_kwh, start_row)

#             # Guardar el archivo
#             wb_out.save(output)
#             output.seek(0)

#             # Botón para descargar el archivo
#             st.download_button(
#                 label="Descargar datos filtrados (Excel)",
#                 data=output,
#                 file_name="datos_filtrados.xlsx",
#                 mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )

#         except Exception as e:
#             st.error(f"Error: {e}")
#         finally:
#             if 'tmp_file_path' in locals():
#                 os.remove(tmp_file_path)

# ----------------------------------------------------

import streamlit as st
import polars as pl
from openpyxl import Workbook
import bcrypt
import io
import os
import tempfile
import re

# Función para verificar el usuario y contraseña
def check_password():
    """Función para autenticar al usuario."""
    def password_entered():
        if bcrypt.checkpw(st.session_state["username"].encode("utf-8"), PASSWORD_CORRECTO):
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # No almacenar la contraseña en la sesión
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.text_input("Usuario", key="username")
        st.text_input("Contraseña", type="password", key="password", on_change=password_entered)
        return False
    
    elif not st.session_state["password_correct"]:
        st.text_input("Usuario", key="username")
        st.text_input("Contraseña", type="password", key="password", on_change=password_entered)
        st.error("Usuario/contraseña incorrectos")
        return False
    else:
        return True
    
# Datos de usuario (¡Importante! Cambia esto por un sistema de almacenamiento seguro)
USUARIO_CORRECTO = "admin"
PASSWORD_CORRECTO = bcrypt.hashpw("admin".encode("utf-8"), bcrypt.gensalt())

# Interfaz Principal
st.title("Extracción y filtrado de datos XLSB (Mes/Año)")

if check_password():
    # Instrucciones principales
    st.header("Cargar archivo XLSB")
    st.write(
        "Por favor, carga un archivo XLSB para extraer y filtrar los datos por mes y año."
        " Asegúrate de que el archivo contenga la hoja correcta con las columnas necesarias."
    )
    uploaded_file = st.file_uploader("Cargar archivo XLSB", type="xlsb")

    if uploaded_file is not None:
        try:
            # Configuración fija
            HOJA_FIJA = 'Libro Fac.Emitidas MM 2025'
            COLUMNAS_FIJAS = ['CUPS', 'Tipo factura', 'Fecha factura', 'Fecha inicial', 'Fecha final', 'Num. días', 'Tarifa', 'Total término variable (kWh)']

            with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
                tmp_file.write(uploaded_file.read())
                tmp_file_path = tmp_file.name

            # Leer el archivo y cargar los datos
            df = pl.read_excel(tmp_file_path, sheet_name=HOJA_FIJA).select(COLUMNAS_FIJAS)
            df = df.drop_nulls()

            # No filtramos por Año-Mes, acumulamos todos los datos
            df_filtered = df.filter(pl.col("Tipo factura") == "Normal")

            # Mostrar los datos filtrados
            st.header("Datos filtrados")
            st.dataframe(df_filtered.head())


            # Generar tablas pivoteadas acumulando por mes
            df_grouped = df_filtered.with_columns(
                pl.col('Fecha inicial').dt.strftime('%Y-%m').alias('Mes')
            ).group_by(['Tarifa', 'Mes']).agg([
                pl.col('Num. días').sum().alias('Suma de num. dias'),
                pl.col('CUPS').count().alias('Cuenta de CUPS'),
                pl.col('Total término variable (kWh)').sum().alias('Suma de kWh')
            ])

            df_pivot_num_dias = df_grouped.pivot(values='Suma de num. dias', index='Tarifa', on='Mes', sort_columns=True)
            df_pivot_cuenta_cups = df_grouped.pivot(values='Cuenta de CUPS', index='Tarifa', on='Mes', sort_columns=True)
            df_pivot_suma_kwh = df_grouped.pivot(values='Suma de kWh', index='Tarifa', on='Mes', sort_columns=True)

            # Ordenar las filas usando la función personalizada
            df_pivot_num_dias = df_pivot_num_dias.sort(by=pl.col("Tarifa"))
            df_pivot_cuenta_cups = df_pivot_cuenta_cups.sort(by=pl.col("Tarifa"))
            df_pivot_suma_kwh = df_pivot_suma_kwh.sort(by=pl.col("Tarifa"))

            # Mostrar los datos filtrados
            st.header("Tablas Generadas. Generando fichero Excel...")
                        
            # Guardar los datos en Excel
            output = io.BytesIO()
            wb_out = Workbook()

            # Hoja 'raw_data' con los datos filtrados
            ws_raw_data = wb_out.active
            ws_raw_data.title = "raw_data"
            ws_raw_data.append(list(df_filtered.columns))
            for row in df_filtered.to_numpy():
                ws_raw_data.append(list(row))

            # Hoja 'consumos' con las tablas pivot
            ws_consumos = wb_out.create_sheet(title="consumos")

            def write_dataframe(ws, df, start_row):
                """Función auxiliar para escribir un DataFrame en la hoja de Excel."""
                ws.append(list(df.columns))  # Escribe los encabezados de las columnas
                for row in df.to_pandas().itertuples(index=False):
                    ws.append(row)
                return start_row + len(df) + 3  # Deja una fila vacía

            start_row = 1
            ws_consumos.append(["Tabla: Suma de num. dias"])
            start_row = write_dataframe(ws_consumos, df_pivot_num_dias, start_row)

            ws_consumos.append(["Tabla: Cuenta de CUPS"])
            start_row = write_dataframe(ws_consumos, df_pivot_cuenta_cups, start_row)

            ws_consumos.append(["Tabla: Suma de kWh"])
            start_row = write_dataframe(ws_consumos, df_pivot_suma_kwh, start_row)

            # Guardar el archivo
            wb_out.save(output)
            output.seek(0)

            # Botón para descargar el archivo
            st.download_button(
                label="Descargar datos filtrados",
                data=output,
                file_name="datos_filtrados.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except Exception as e:
            st.error(f"Error: {e}")
        finally:
            if 'tmp_file_path' in locals():
                os.remove(tmp_file_path)
